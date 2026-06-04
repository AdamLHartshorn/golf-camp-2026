import json, re
from pathlib import Path
from openpyxl import load_workbook

path=Path('/Users/adamlhartshorn/Downloads/2026 Golf Trip Attendees (1).xlsx')
wb=load_workbook(path, data_only=True)
ws=wb['Attendee Scores and Statements']
questions_cell=str(ws.cell(1,3).value or '')
q_pattern=re.compile(r'(\d+)\.\s*(.*?)(?=\n\d+\.\s*|\Z)', re.S)
questions=[]
for num, q in q_pattern.findall(questions_cell):
    questions.append({'number': int(num), 'question': ' '.join(q.strip().split())})

def split_answers(text):
    text=str(text or '')
    matches=list(re.finditer(r'(?:^|\n)\s*(\d+)\.\s*', text))
    answers={}
    for i,m in enumerate(matches):
        num=int(m.group(1))
        start=m.end()
        end=matches[i+1].start() if i+1<len(matches) else len(text)
        answer=text[start:end].strip()
        answers[num]=answer
    return answers

rows=[]
for r in range(2, ws.max_row+1):
    name=ws.cell(r,1).value
    text=ws.cell(r,3).value
    if not name or not str(name).strip() or not text or not str(text).strip():
        continue
    answers=split_answers(text)
    rows.append({'workbook_name': str(name).strip(), 'answers_by_number': answers})

print(json.dumps({'questions': questions, 'rows': rows}, ensure_ascii=False, indent=2))
